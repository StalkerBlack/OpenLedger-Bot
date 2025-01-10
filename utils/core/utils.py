import asyncio
import random
import openpyxl

from .logger import logger


async def smart_sleep(up, to):
    duration = random.randint(up, to)
    logger.info(f"💤 Waiting {duration:.2f} seconds.")
    await asyncio.sleep(duration)

def get_accounts_data():
    try:
        book = openpyxl.load_workbook('data/accounts.xlsx', read_only=True)
        sheet = book.active
    except FileNotFoundError:
        logger.error('The file accounts.xlsx was not found')
        return [], [], [], []
    except Exception as e:
        logger.error(f"Error when opening a file: {e}")
        return [], [], [], []

    accounts = []
    
    for row in range(2, sheet.max_row + 1):
        try:
            account_name = sheet.cell(row=row, column=1).value
            proxies = sheet.cell(row=row, column=2).value
            token = sheet.cell(row=row, column=3).value

            if not any([account_name, proxies, token]):
                continue

            if not all([account_name, proxies, token]):
                logger.warning(f"String {row}: Insufficient data for account")
                continue

            account = {
                'account_name': account_name,
                'proxies': proxies,
                'token': token
            }
            accounts.append(account)

        except Exception as e:
            logger.error(f"Error when reading a string {row}: {e}")

    return accounts